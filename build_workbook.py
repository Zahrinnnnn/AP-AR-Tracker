"""
AP/AR Tracker - Workbook Builder
Author: Zahrin Bin Jasni

Builds the full Excel workbook for the AP/AR Tracker project.
Run this script to regenerate the workbook from scratch.

Phases:
    1 - Workbook structure, Settings, Lookup Tables
    2 - AR Ledger and AP Ledger with all formulas
    3 - AR and AP Aging Reports
    4 - Dashboard KPIs and charts
    5 - Conditional formatting, data validation, print layout
"""

import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ─────────────────────────────────────────────────────────────────
# OUTPUT
# ─────────────────────────────────────────────────────────────────

OUTPUT_FILE = "AP_AR_Tracker.xlsx"

# ─────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────

NAVY        = "1F3864"
NAVY_LIGHT  = "2E4D7B"
WHITE       = "FFFFFF"
YELLOW_IN   = "FFFFC6"   # input cells
GREY_FORM   = "D9D9D9"   # formula cells
BLUE_ALT    = "DCE6F1"   # alternating row tint
RED_DARK    = "C00000"   # 90+ overdue
RED_LIGHT   = "FFCCCC"
AMBER_DARK  = "C07000"   # 61-90 overdue
AMBER_LIGHT = "FFE0B2"
YELLOW_DARK = "7D6608"   # 31-60 overdue
YELLOW_LIGHT= "FFF9C4"
GREEN_DARK  = "375623"
GREEN_LIGHT = "E2EFDA"   # current / paid
GREY_LIGHT  = "F2F2F2"

# ─────────────────────────────────────────────────────────────────
# NUMBER FORMATS
# ─────────────────────────────────────────────────────────────────

FMT_CURRENCY = '#,##0.00'
FMT_DATE     = 'DD/MM/YYYY'
FMT_INT      = '#,##0'
FMT_PERCENT  = '0.0%'

# ─────────────────────────────────────────────────────────────────
# LEDGER COLUMN POSITIONS  (1-based, shared by AR and AP Ledger)
# ─────────────────────────────────────────────────────────────────

COL_NUM      = 1   # Invoice / Bill Number
COL_DATE     = 2   # Invoice / Bill Date
COL_DUE      = 3   # Due Date
COL_ENTITY   = 4   # Client / Vendor Name
COL_DESC     = 5   # Description
COL_AMOUNT   = 6   # Amount (excl. Tax)
COL_TAX      = 7   # Tax Amount          <- formula
COL_TOTAL    = 8   # Total Amount        <- formula
COL_PAY_DATE = 9   # Payment Date
COL_PAY_AMT  = 10  # Payment Amount
COL_BALANCE  = 11  # Outstanding Balance <- formula
COL_DAYS     = 12  # Days Overdue        <- formula
COL_AGING    = 13  # Aging Bucket        <- formula
COL_STATUS   = 14  # Status              <- formula
COL_NOTES    = 15  # Notes

LEDGER_HEADER_ROW = 2   # row 1 is the sheet title
LEDGER_DATA_START = 3   # data begins here
LEDGER_SAMPLE_END = 7   # last sample row (5 rows of sample data)

# ─────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────

def thin_border():
    """Thin border on all four sides."""
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def header_style(cell, text):
    """Navy background, white bold text — used for column headers."""
    cell.value = text
    cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()


def section_header_style(cell, text):
    """Slightly lighter navy for sub-section labels."""
    cell.value = text
    cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill      = PatternFill("solid", fgColor=NAVY_LIGHT)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def input_style(cell, value=None, fmt=None):
    """Light yellow — signals this cell accepts manual input."""
    if value is not None:
        cell.value = value
    cell.fill      = PatternFill("solid", fgColor=YELLOW_IN)
    cell.font      = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt


def formula_style(cell, formula=None, fmt=None):
    """Light grey — signals this cell is calculated, not for direct editing."""
    if formula is not None:
        cell.value = formula
    cell.fill      = PatternFill("solid", fgColor=GREY_FORM)
    cell.font      = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt


def set_col_widths(ws, widths: dict):
    """Set column widths. widths = {"A": 20, "B": 15, ...}"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def title_style(cell, text, size=14):
    """Large navy bold title at the top of a sheet."""
    cell.value = text
    cell.font  = Font(name="Calibri", bold=True, size=size, color=NAVY)


# ─────────────────────────────────────────────────────────────────
# PHASE 1A  —  WORKBOOK STRUCTURE
# ─────────────────────────────────────────────────────────────────

def create_workbook():
    """
    Create a fresh workbook with all 7 sheets in the correct tab order.
    The default 'Sheet' created by openpyxl is renamed to the first tab.
    """
    wb = openpyxl.Workbook()

    sheet_names = [
        "Dashboard",
        "AR Ledger",
        "AP Ledger",
        "AR Aging Report",
        "AP Aging Report",
        "Settings",
        "Lookup Tables",
    ]

    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(title=name)

    print("  [ok] 7 sheets created")
    return wb


# ─────────────────────────────────────────────────────────────────
# PHASE 1B  —  SETTINGS SHEET
# ─────────────────────────────────────────────────────────────────

def build_settings(wb):
    """
    Settings sheet — company-level configuration referenced by formulas
    across all other sheets. All value cells are yellow input cells.

    Row positions (important — formulas elsewhere reference these):
        B4  = Company Name
        B5  = Report Date
        B6  = Tax Rate (%)
        B7  = Currency
        B8  = Aging Threshold 1   (default 30)
        B9  = Aging Threshold 2   (default 60)
        B10 = Aging Threshold 3   (default 90)
        B11 = AR Invoice Prefix
        B12 = AP Bill Prefix
    """
    ws = wb["Settings"]

    title_style(ws["A1"], "AP/AR Tracker — Settings")
    ws.row_dimensions[1].height = 28

    header_style(ws["A3"], "Setting")
    header_style(ws["B3"], "Value")
    header_style(ws["C3"], "Notes")

    rows = [
        ("Company Name",       "Your Company Sdn Bhd",   "Shown in all report headers"),
        ("Report Date",        "=TODAY()",                "Auto-set to today; override if needed"),
        ("Tax Rate (%)",       6,                         "SST rate applied to invoice amounts"),
        ("Currency",           "RM",                      "Display currency symbol"),
        ("Aging Threshold 1",  30,                        "Upper limit for first overdue bucket"),
        ("Aging Threshold 2",  60,                        "Upper limit for second overdue bucket"),
        ("Aging Threshold 3",  90,                        "Upper limit for third overdue bucket"),
        ("AR Invoice Prefix",  "INV-",                    "Auto-number prefix for AR invoices"),
        ("AP Bill Prefix",     "BILL-",                   "Auto-number prefix for AP bills"),
    ]

    for i, (label, value, note) in enumerate(rows, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font   = Font(name="Calibri", bold=True, size=10)
        label_cell.border = thin_border()

        input_style(ws.cell(row=i, column=2), value)

        note_cell = ws.cell(row=i, column=3, value=note)
        note_cell.font   = Font(name="Calibri", size=10, italic=True, color="595959")
        note_cell.border = thin_border()

    # Format the Report Date cell as a date
    ws["B5"].number_format = FMT_DATE

    set_col_widths(ws, {"A": 22, "B": 28, "C": 48})
    print("  [ok] Settings sheet built")


# ─────────────────────────────────────────────────────────────────
# PHASE 1C  —  LOOKUP TABLES SHEET
# ─────────────────────────────────────────────────────────────────

def build_lookup_tables(wb):
    """
    Lookup Tables sheet — reference lists used in dropdowns and XLOOKUP
    formulas across the ledger sheets.

    Four tables: Clients, Vendors, Payment Terms, Account Codes.
    Client names live in B5:B9  — referenced by AR Ledger dropdown.
    Vendor names live in E5:E9  — referenced by AP Ledger dropdown.
    """
    ws = wb["Lookup Tables"]

    title_style(ws["A1"], "Lookup Tables — Reference Data")
    ws.row_dimensions[1].height = 28

    # ── CLIENT LIST ──────────────────────────────────────────────
    section_header_style(ws["A3"], "  Client List")
    ws.merge_cells("A3:B3")
    header_style(ws["A4"], "Client ID")
    header_style(ws["B4"], "Client Name")

    clients = [
        ("CL-001", "Axiata Group Berhad"),
        ("CL-002", "Tenaga Nasional Berhad"),
        ("CL-003", "CIMB Group Holdings"),
        ("CL-004", "Petronas Dagangan"),
        ("CL-005", "Top Glove Corporation"),
    ]
    for i, (cid, name) in enumerate(clients, start=5):
        input_style(ws.cell(row=i, column=1), cid)
        input_style(ws.cell(row=i, column=2), name)

    # ── VENDOR LIST ───────────────────────────────────────────────
    section_header_style(ws["D3"], "  Vendor List")
    ws.merge_cells("D3:E3")
    header_style(ws["D4"], "Vendor ID")
    header_style(ws["E4"], "Vendor Name")

    vendors = [
        ("VD-001", "Pos Malaysia Berhad"),
        ("VD-002", "Telekom Malaysia"),
        ("VD-003", "Digi Telecommunications"),
        ("VD-004", "Syabas Water Supply"),
        ("VD-005", "TNB (Electricity)"),
    ]
    for i, (vid, name) in enumerate(vendors, start=5):
        input_style(ws.cell(row=i, column=4), vid)
        input_style(ws.cell(row=i, column=5), name)

    # ── PAYMENT TERMS ─────────────────────────────────────────────
    section_header_style(ws["G3"], "  Payment Terms")
    ws.merge_cells("G3:H3")
    header_style(ws["G4"], "Term Code")
    header_style(ws["H4"], "Description")

    terms = [
        ("NET7",  "Payment due in 7 days"),
        ("NET14", "Payment due in 14 days"),
        ("NET30", "Payment due in 30 days"),
        ("NET60", "Payment due in 60 days"),
        ("COD",   "Cash on delivery"),
        ("EOM",   "End of month"),
    ]
    for i, (code, desc) in enumerate(terms, start=5):
        input_style(ws.cell(row=i, column=7), code)
        input_style(ws.cell(row=i, column=8), desc)

    # ── ACCOUNT CODES ─────────────────────────────────────────────
    section_header_style(ws["J3"], "  Account Codes")
    ws.merge_cells("J3:K3")
    header_style(ws["J4"], "Code")
    header_style(ws["K4"], "Account Name")

    accounts = [
        ("1100", "Accounts Receivable"),
        ("2100", "Accounts Payable"),
        ("4000", "Sales Revenue"),
        ("5000", "Cost of Goods Sold"),
        ("6000", "Operating Expenses"),
        ("2200", "SST Payable"),
    ]
    for i, (code, name) in enumerate(accounts, start=5):
        input_style(ws.cell(row=i, column=10), code)
        input_style(ws.cell(row=i, column=11), name)

    set_col_widths(ws, {
        "A": 10, "B": 26, "C": 3,
        "D": 10, "E": 26, "F": 3,
        "G": 10, "H": 28, "I": 3,
        "J": 9,  "K": 26,
    })
    print("  [ok] Lookup Tables sheet built")


# ─────────────────────────────────────────────────────────────────
# PHASE 2  —  LEDGER FORMULAS  (shared logic for AR and AP)
# ─────────────────────────────────────────────────────────────────

def get_ledger_formulas(row):
    """
    Returns the formula string for each calculated column at the given row.
    Thresholds are read from the Settings sheet so the user can change them
    without touching any formula.
    """
    r = row

    tax_amount = f"=F{r}*Settings!$B$6/100"

    total_amount = f"=F{r}+G{r}"

    outstanding = f"=H{r}-J{r}"

    days_overdue = f"=IF(K{r}<=0,0,MAX(0,TODAY()-C{r}))"

    # Nested IF instead of IFS() — more compatible across Excel versions
    # and avoids openpyxl XML serialisation issues with IFS().
    aging_bucket = (
        f'=IF(K{r}<=0,"Paid",'
        f'IF(L{r}=0,"Current",'
        f'IF(L{r}<=Settings!$B$8,"1-30 Days",'
        f'IF(L{r}<=Settings!$B$9,"31-60 Days",'
        f'IF(L{r}<=Settings!$B$10,"61-90 Days","90+ Days")))))'
    )

    status = (
        f'=IF(K{r}<=0,"Paid",'
        f'IF(J{r}>0,"Partial",'
        f'IF(L{r}>0,"Overdue","Outstanding")))'
    )

    return {
        COL_TAX:     (tax_amount,    FMT_CURRENCY),
        COL_TOTAL:   (total_amount,  FMT_CURRENCY),
        COL_BALANCE: (outstanding,   FMT_CURRENCY),
        COL_DAYS:    (days_overdue,  FMT_INT),
        COL_AGING:   (aging_bucket,  None),
        COL_STATUS:  (status,        None),
    }


def write_ledger_row(ws, row, data_tuple):
    """
    Writes one row of data into the ledger sheet and applies cell styles.
    data_tuple columns:
        0  Invoice/Bill Number
        1  Invoice/Bill Date
        2  Due Date
        3  Client/Vendor Name
        4  Description
        5  Amount (excl. Tax)
        6  (Tax — formula, skip)
        7  (Total — formula, skip)
        8  Payment Date  (None if unpaid)
        9  Payment Amount
        10 (Balance — formula, skip)
        11 (Days — formula, skip)
        12 (Aging — formula, skip)
        13 (Status — formula, skip)
        14 Notes
    """
    num, inv_date, due_date, entity, desc, amount, _, _, pay_date, pay_amt, _, _, _, _, notes = data_tuple

    # Input cells
    input_style(ws.cell(row=row, column=COL_NUM),    num)
    input_style(ws.cell(row=row, column=COL_DATE),   inv_date,  FMT_DATE)
    input_style(ws.cell(row=row, column=COL_DUE),    due_date,  FMT_DATE)
    input_style(ws.cell(row=row, column=COL_ENTITY), entity)
    input_style(ws.cell(row=row, column=COL_DESC),   desc)
    input_style(ws.cell(row=row, column=COL_AMOUNT), amount,    FMT_CURRENCY)

    if pay_date:
        input_style(ws.cell(row=row, column=COL_PAY_DATE), pay_date, FMT_DATE)
    else:
        input_style(ws.cell(row=row, column=COL_PAY_DATE), "")

    input_style(ws.cell(row=row, column=COL_PAY_AMT), pay_amt if pay_amt else 0, FMT_CURRENCY)
    input_style(ws.cell(row=row, column=COL_NOTES),   notes)

    # Formula cells
    formulas = get_ledger_formulas(row)
    for col, (formula, fmt) in formulas.items():
        formula_style(ws.cell(row=row, column=col), formula, fmt)


def build_ledger_sheet(wb, sheet_name, table_name, headers, sample_data):
    """
    Generic function that builds either the AR Ledger or AP Ledger sheet.
    Creates the title, column headers, sample data rows, and an Excel Table.
    """
    ws = wb[sheet_name]

    # Sheet title row
    ws.row_dimensions[1].height = 28
    title_style(ws["A1"], f"{sheet_name} — Invoice & Payment Records")
    ws.merge_cells("A1:O1")

    # Column header row
    ws.row_dimensions[LEDGER_HEADER_ROW].height = 36
    for col_index, heading in enumerate(headers, start=1):
        header_style(ws.cell(row=LEDGER_HEADER_ROW, column=col_index), heading)

    # Sample data rows
    for i, row_data in enumerate(sample_data):
        row_num = LEDGER_DATA_START + i
        ws.row_dimensions[row_num].height = 18
        write_ledger_row(ws, row_num, row_data)

    # Register as an Excel Table so formulas auto-fill on new rows
    last_row = LEDGER_DATA_START + len(sample_data) - 1
    table_ref = f"A{LEDGER_HEADER_ROW}:{get_column_letter(COL_NOTES)}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)

    # Freeze top 2 rows and first column
    ws.freeze_panes = "B3"

    # Column widths
    set_col_widths(ws, {
        "A": 13, "B": 13, "C": 13, "D": 26, "E": 28,
        "F": 16, "G": 14, "H": 14, "I": 14, "J": 16,
        "K": 18, "L": 13, "M": 14, "N": 13, "O": 22,
    })

    ws.sheet_view.showGridLines = True
    print(f"  [ok] {sheet_name} built ({len(sample_data)} sample rows)")


# ─────────────────────────────────────────────────────────────────
# PHASE 2  —  AR LEDGER
# ─────────────────────────────────────────────────────────────────

AR_HEADERS = [
    "Invoice Number", "Invoice Date", "Due Date", "Client Name",
    "Description", "Amount (excl. Tax)", "Tax Amount", "Total Amount",
    "Payment Date", "Payment Amount", "Outstanding Balance",
    "Days Overdue", "Aging Bucket", "Status", "Notes",
]

# Five sample invoices spanning all aging buckets (as of 2026-03-29)
AR_SAMPLE = [
    # Paid — full payment received
    ("INV-001", datetime.date(2026, 1,  5), datetime.date(2026, 1, 15),
     "Axiata Group Berhad",    "Software Development",  5000.00,
     None, None,
     datetime.date(2026, 1, 20), 5300.00,
     None, None, None, None, ""),

    # Partial payment — 53 days overdue → 31-60 Days bucket
    ("INV-002", datetime.date(2026, 1, 20), datetime.date(2026, 2,  4),
     "Tenaga Nasional Berhad", "Consulting Services",   8000.00,
     None, None,
     datetime.date(2026, 2, 20), 4000.00,
     None, None, None, None, "Follow up required"),

    # No payment — 32 days overdue → 31-60 Days bucket
    ("INV-003", datetime.date(2026, 2, 10), datetime.date(2026, 2, 25),
     "CIMB Group Holdings",    "Staff Training",        3500.00,
     None, None,
     None, 0,
     None, None, None, None, ""),

    # No payment — 14 days overdue → 1-30 Days bucket
    ("INV-004", datetime.date(2026, 3,  1), datetime.date(2026, 3, 15),
     "Axiata Group Berhad",    "System Maintenance",    2000.00,
     None, None,
     None, 0,
     None, None, None, None, ""),

    # No payment — due 2026-04-19, not yet overdue → Current
    ("INV-005", datetime.date(2026, 3, 20), datetime.date(2026, 4, 19),
     "Petronas Dagangan",      "Software License",     12000.00,
     None, None,
     None, 0,
     None, None, None, None, ""),
]


def build_ar_ledger(wb):
    build_ledger_sheet(wb, "AR Ledger", "AR_Ledger", AR_HEADERS, AR_SAMPLE)


# ─────────────────────────────────────────────────────────────────
# PHASE 2  —  AP LEDGER
# ─────────────────────────────────────────────────────────────────

AP_HEADERS = [
    "Bill Number", "Bill Date", "Due Date", "Vendor Name",
    "Description", "Amount (excl. Tax)", "Tax Amount", "Total Amount",
    "Payment Date", "Payment Amount", "Outstanding Balance",
    "Days Overdue", "Aging Bucket", "Status", "Notes",
]

# Five sample vendor bills
AP_SAMPLE = [
    # Paid
    ("BILL-001", datetime.date(2026, 1,  8), datetime.date(2026, 1, 18),
     "Pos Malaysia Berhad",     "Courier Services",      500.00,
     None, None,
     datetime.date(2026, 1, 20), 530.00,
     None, None, None, None, ""),

    # Paid
    ("BILL-002", datetime.date(2026, 1, 25), datetime.date(2026, 2,  9),
     "Telekom Malaysia",        "Internet Subscription", 350.00,
     None, None,
     datetime.date(2026, 2, 12), 371.00,
     None, None, None, None, ""),

    # No payment — 35 days overdue → 31-60 Days
    ("BILL-003", datetime.date(2026, 2,  8), datetime.date(2026, 2, 22),
     "Digi Telecommunications", "Mobile Services",       200.00,
     None, None,
     None, 0,
     None, None, None, None, ""),

    # No payment — 10 days overdue → 1-30 Days
    ("BILL-004", datetime.date(2026, 3,  5), datetime.date(2026, 3, 19),
     "Syabas Water Supply",     "Monthly Water Bill",    150.00,
     None, None,
     None, 0,
     None, None, None, None, ""),

    # Not yet due → Current
    ("BILL-005", datetime.date(2026, 3, 15), datetime.date(2026, 4, 14),
     "TNB (Electricity)",       "Electricity Bill",      800.00,
     None, None,
     None, 0,
     None, None, None, None, ""),
]


def build_ap_ledger(wb):
    build_ledger_sheet(wb, "AP Ledger", "AP_Ledger", AP_HEADERS, AP_SAMPLE)


# ─────────────────────────────────────────────────────────────────
# PHASE 3  —  AGING REPORT  (shared logic for AR and AP)
# ─────────────────────────────────────────────────────────────────

AGING_BUCKETS = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days"]


def build_aging_report(wb, sheet_name, ledger_sheet, entity_col_letter,
                       entity_names, include_next_due=False):
    """
    Builds one aging report (AR or AP).

    sheet_name        : "AR Aging Report" or "AP Aging Report"
    ledger_sheet      : "'AR Ledger'" or "'AP Ledger'" (with quotes for formula use)
    entity_col_letter : column letter for client/vendor in the ledger (D)
    entity_names      : list of client or vendor names to populate rows
    include_next_due  : AP report has an extra "Next Payment Due" column
    """
    ws = wb[sheet_name]

    title_style(ws["A1"], f"{sheet_name} — Outstanding Balance Summary")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    # Subtitle: report date pulled from Settings
    ws["A2"] = "Report Date:"
    ws["A2"].font = Font(name="Calibri", bold=True, size=10)
    ws["B2"] = "=Settings!$B$5"
    ws["B2"].number_format = FMT_DATE
    ws["B2"].font = Font(name="Calibri", size=10)

    # Column headers — row 4
    col_headers = ["Entity Name", "Total Outstanding",
                   "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days"]
    if include_next_due:
        col_headers.append("Next Payment Due")

    header_row = 4
    ws.row_dimensions[header_row].height = 32
    for col_index, heading in enumerate(col_headers, start=1):
        header_style(ws.cell(row=header_row, column=col_index), heading)

    # Helper: SUMIFS formula for a specific aging bucket
    def sumifs_bucket(entity_cell_ref, bucket_label):
        balance_col  = f"{ledger_sheet}!$K:$K"
        entity_col   = f"{ledger_sheet}!${entity_col_letter}:${entity_col_letter}"
        aging_col    = f"{ledger_sheet}!$M:$M"
        return (
            f'=SUMIFS({balance_col},'
            f'{entity_col},{entity_cell_ref},'
            f'{aging_col},"{bucket_label}")'
        )

    # Helper: SUMIF for total outstanding (all non-paid buckets)
    def sumif_total(entity_cell_ref):
        balance_col = f"{ledger_sheet}!$K:$K"
        entity_col  = f"{ledger_sheet}!${entity_col_letter}:${entity_col_letter}"
        return f'=SUMIF({entity_col},{entity_cell_ref},{balance_col})'

    # Helper: MINIFS for next payment due date (AP only)
    def minifs_next_due(entity_cell_ref):
        due_col    = f"{ledger_sheet}!$C:$C"
        entity_col = f"{ledger_sheet}!${entity_col_letter}:${entity_col_letter}"
        status_col = f"{ledger_sheet}!$N:$N"
        return (
            f'=IFERROR(MINIFS({due_col},{entity_col},{entity_cell_ref},'
            f'{status_col},"<>Paid"),"")'
        )

    # Entity rows
    data_start = header_row + 1
    for i, entity_name in enumerate(entity_names):
        row = data_start + i
        entity_ref = f"A{row}"

        ws.row_dimensions[row].height = 18

        # Column A: entity name (linked from Lookup Tables for easy maintenance)
        name_cell = ws.cell(row=row, column=1, value=entity_name)
        name_cell.font   = Font(name="Calibri", bold=True, size=10)
        name_cell.border = thin_border()

        # Column B: total outstanding
        formula_style(ws.cell(row=row, column=2), sumif_total(entity_ref), FMT_CURRENCY)

        # Columns C-G: one per aging bucket
        for bucket_col, bucket in enumerate(AGING_BUCKETS, start=3):
            formula_style(ws.cell(row=row, column=bucket_col),
                          sumifs_bucket(entity_ref, bucket), FMT_CURRENCY)

        # Column H (AP only): next payment due
        if include_next_due:
            formula_style(ws.cell(row=row, column=8),
                          minifs_next_due(entity_ref), FMT_DATE)

    # Totals row
    totals_row = data_start + len(entity_names)
    ws.row_dimensions[totals_row].height = 20

    totals_label = ws.cell(row=totals_row, column=1, value="TOTAL")
    totals_label.font   = Font(name="Calibri", bold=True, size=11, color=WHITE)
    totals_label.fill   = PatternFill("solid", fgColor=NAVY)
    totals_label.border = thin_border()

    last_data_row = totals_row - 1
    for col in range(2, len(col_headers) + 1):
        col_letter = get_column_letter(col)
        total_formula = f"=SUM({col_letter}{data_start}:{col_letter}{last_data_row})"
        total_cell = ws.cell(row=totals_row, column=col)
        total_cell.value          = total_formula
        total_cell.font           = Font(name="Calibri", bold=True, color=WHITE)
        total_cell.fill           = PatternFill("solid", fgColor=NAVY)
        total_cell.border         = thin_border()
        total_cell.number_format  = FMT_CURRENCY

    # Column widths
    set_col_widths(ws, {
        "A": 28, "B": 18, "C": 14,
        "D": 14, "E": 14, "F": 14, "G": 14, "H": 18,
    })

    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False

    print(f"  [ok] {sheet_name} built")


AR_AGING_CLIENTS = [
    "Axiata Group Berhad",
    "Tenaga Nasional Berhad",
    "CIMB Group Holdings",
    "Petronas Dagangan",
    "Top Glove Corporation",
]

AP_AGING_VENDORS = [
    "Pos Malaysia Berhad",
    "Telekom Malaysia",
    "Digi Telecommunications",
    "Syabas Water Supply",
    "TNB (Electricity)",
]


def build_ar_aging_report(wb):
    build_aging_report(
        wb,
        sheet_name     = "AR Aging Report",
        ledger_sheet   = "'AR Ledger'",
        entity_col_letter = "D",
        entity_names   = AR_AGING_CLIENTS,
        include_next_due = False,
    )


def build_ap_aging_report(wb):
    build_aging_report(
        wb,
        sheet_name     = "AP Aging Report",
        ledger_sheet   = "'AP Ledger'",
        entity_col_letter = "D",
        entity_names   = AP_AGING_VENDORS,
        include_next_due = True,
    )


# ─────────────────────────────────────────────────────────────────
# PHASE 4  —  DASHBOARD
# ─────────────────────────────────────────────────────────────────

def write_kpi_card(ws, label_row, value_row, col_start, col_end, label, formula, fmt=FMT_CURRENCY, value_size=20):
    """
    Writes a KPI card using two merged cell rows: label on top, value below.
    Uses navy for label and a large bold number for the value.
    """
    ws.merge_cells(
        start_row=label_row, start_column=col_start,
        end_row=label_row,   end_column=col_end
    )
    label_cell = ws.cell(row=label_row, column=col_start, value=label)
    label_cell.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
    label_cell.fill      = PatternFill("solid", fgColor=NAVY)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border    = thin_border()

    ws.merge_cells(
        start_row=value_row, start_column=col_start,
        end_row=value_row,   end_column=col_end
    )
    value_cell = ws.cell(row=value_row, column=col_start, value=formula)
    value_cell.font          = Font(name="Calibri", bold=True, size=value_size, color=NAVY)
    value_cell.fill          = PatternFill("solid", fgColor=BLUE_ALT)
    value_cell.alignment     = Alignment(horizontal="center", vertical="center")
    value_cell.border        = thin_border()
    value_cell.number_format = fmt


def build_dashboard(wb):
    """
    Dashboard sheet — KPI cards, summary, and charts.
    No manual input: all values driven by formulas from the ledger sheets.

    Layout:
        Row 1      : Main title
        Row 2      : Company name + report date
        Row 3      : Blank separator
        Row 4-5    : KPI row 1 (3 cards)
        Row 6-7    : KPI row 2 (3 cards)
        Row 8      : Blank separator
        Row 10-30  : Charts (AR Aging + AP Aging side by side)
        Row 32-50  : Charts (Top 5 Clients + Top 5 Vendors)
        Row 55+    : Chart data tables (used as chart sources)
    """
    ws = wb["Dashboard"]
    ws.sheet_view.showGridLines = False

    # ── TITLE ────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"].value          = "AP / AR TRACKER  —  DASHBOARD"
    ws["A1"].font           = Font(name="Calibri", bold=True, size=18, color=WHITE)
    ws["A1"].fill           = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment      = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # ── COMPANY & DATE ───────────────────────────────────────────
    ws["A2"].value          = "=Settings!$B$4"
    ws["A2"].font           = Font(name="Calibri", bold=True, size=11, color=NAVY)
    ws["A2"].alignment      = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:F2")

    ws["J2"].value          = "Report Date:"
    ws["J2"].font           = Font(name="Calibri", bold=True, size=10, color="595959")
    ws["K2"].value          = "=Settings!$B$5"
    ws["K2"].number_format  = FMT_DATE
    ws["K2"].font           = Font(name="Calibri", size=10)
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 8  # visual gap

    # ── KPI ROW 1 (rows 4–5) ─────────────────────────────────────
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 44

    write_kpi_card(ws, 4, 5, 1, 4,
                   "Total AR Outstanding (RM)",
                   "=SUMIF('AR Ledger'!$N$3:$N$1000,\"<>Paid\",'AR Ledger'!$K$3:$K$1000)")

    write_kpi_card(ws, 4, 5, 5, 8,
                   "Total AP Outstanding (RM)",
                   "=SUMIF('AP Ledger'!$N$3:$N$1000,\"<>Paid\",'AP Ledger'!$K$3:$K$1000)")

    write_kpi_card(ws, 4, 5, 9, 12,
                   "Net Position: AR minus AP (RM)",
                   "=SUMIF('AR Ledger'!$N$3:$N$1000,\"<>Paid\",'AR Ledger'!$K$3:$K$1000)"
                   "-SUMIF('AP Ledger'!$N$3:$N$1000,\"<>Paid\",'AP Ledger'!$K$3:$K$1000)")

    # ── KPI ROW 2 (rows 6–7) ─────────────────────────────────────
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 44

    write_kpi_card(ws, 6, 7, 1, 4,
                   "Total AR Overdue 90+ Days (RM)",
                   "=SUMIF('AR Ledger'!$M$3:$M$1000,\"90+ Days\",'AR Ledger'!$K$3:$K$1000)")

    write_kpi_card(ws, 6, 7, 5, 8,
                   "Number of Overdue AR Invoices",
                   "=COUNTIF('AR Ledger'!$N$3:$N$1000,\"Overdue\")"
                   "+COUNTIF('AR Ledger'!$N$3:$N$1000,\"Partial\")",
                   fmt=FMT_INT, value_size=20)

    write_kpi_card(ws, 6, 7, 9, 12,
                   "Collection Rate (Paid / Total Invoiced)",
                   "=IFERROR(1-SUMIF('AR Ledger'!$N$3:$N$1000,\"<>Paid\",'AR Ledger'!$K$3:$K$1000)"
                   "/SUM('AR Ledger'!$H$3:$H$1000),0)",
                   fmt=FMT_PERCENT, value_size=20)

    ws.row_dimensions[8].height = 10  # gap before charts

    # ── CHART DATA TABLES (rows 45–57) ───────────────────────────
    #
    # Each table is vertical: 5 rows x 1 value column.
    # Vertical layout = 1 series with 5 data points (one per bucket).
    # Horizontal layout (1 row x 5 cols) = 5 series with 1 point each
    # which produces the "Series1-5 with empty bars" bug.
    #
    # Formulas use direct SUMIF against the ledger — same pattern as
    # the KPI cards above, which are confirmed working. Avoids an
    # extra formula hop through the Aging Report sheets.
    #
    # AR chart data : rows 46-50  (header at 45)
    # AP chart data : rows 53-57  (header at 52)

    ar_data_start = 45
    ap_data_start = 52

    def write_chart_data_table(start_row, ledger_sheet_name):
        """
        Writes a 5-row vertical SUMIF table for one ledger's aging buckets.
        start_row  : row of the header label
        ledger_sheet_name : e.g. 'AR Ledger' (without quotes)
        """
        ws.cell(row=start_row, column=1).value = "Aging Bucket"
        ws.cell(row=start_row, column=2).value = "Outstanding (RM)"

        for i, bucket in enumerate(AGING_BUCKETS):
            r = start_row + 1 + i

            # Bucket label — this becomes the chart category axis label
            ws.cell(row=r, column=1).value = bucket

            # SUMIF: sum all Outstanding Balances in this bucket
            ws.cell(row=r, column=2).value = (
                f"=SUMIF('{ledger_sheet_name}'!$M$3:$M$1000,"
                f'"{bucket}",'
                f"'{ledger_sheet_name}'!$K$3:$K$1000)"
            )
            ws.cell(row=r, column=2).number_format = FMT_CURRENCY

    write_chart_data_table(ar_data_start, "AR Ledger")
    write_chart_data_table(ap_data_start, "AP Ledger")

    # ── AR AGING CHART ───────────────────────────────────────────
    ar_chart = BarChart()
    ar_chart.type         = "bar"       # horizontal bars
    ar_chart.grouping     = "clustered"
    ar_chart.title        = "AR Aging Distribution"
    ar_chart.y_axis.title = "Aging Bucket"
    ar_chart.x_axis.title = "Outstanding (RM)"
    ar_chart.style        = 10
    ar_chart.width        = 16
    ar_chart.height       = 12
    ar_chart.legend       = None        # single series, no legend needed

    # Data: column B, 5 rows (one value per bucket)
    ar_data = Reference(ws,
                        min_col=2, max_col=2,
                        min_row=ar_data_start + 1,
                        max_row=ar_data_start + 5)

    # Categories: column A, same rows (bucket labels)
    ar_cats = Reference(ws,
                        min_col=1,
                        min_row=ar_data_start + 1,
                        max_row=ar_data_start + 5)

    ar_chart.add_data(ar_data, titles_from_data=False)
    ar_chart.set_categories(ar_cats)
    ar_chart.series[0].title = SeriesLabel(v="AR Outstanding (RM)")
    ws.add_chart(ar_chart, "A10")

    # ── AP AGING CHART ───────────────────────────────────────────
    ap_chart = BarChart()
    ap_chart.type         = "bar"
    ap_chart.grouping     = "clustered"
    ap_chart.title        = "AP Aging Distribution"
    ap_chart.y_axis.title = "Aging Bucket"
    ap_chart.x_axis.title = "Outstanding (RM)"
    ap_chart.style        = 10
    ap_chart.width        = 16
    ap_chart.height       = 12
    ap_chart.legend       = None

    ap_data = Reference(ws,
                        min_col=2, max_col=2,
                        min_row=ap_data_start + 1,
                        max_row=ap_data_start + 5)

    ap_cats = Reference(ws,
                        min_col=1,
                        min_row=ap_data_start + 1,
                        max_row=ap_data_start + 5)

    ap_chart.add_data(ap_data, titles_from_data=False)
    ap_chart.set_categories(ap_cats)
    ap_chart.series[0].title = SeriesLabel(v="AP Outstanding (RM)")
    ws.add_chart(ap_chart, "G10")

    # Column widths on dashboard
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14

    print("  [ok] Dashboard built")


# ─────────────────────────────────────────────────────────────────
# PHASE 5A  —  CONDITIONAL FORMATTING
# ─────────────────────────────────────────────────────────────────

def apply_conditional_formatting(wb):
    """
    Applies colour-coded conditional formatting to the Aging Bucket column
    in both ledger sheets and aging reports.

    Rules (applied to Aging Bucket column M in ledgers):
        "90+ Days"  -> red background
        "61-90 Days"-> amber background
        "31-60 Days"-> yellow background
        "1-30 Days" -> light blue background
        "Current"   -> green background
        "Paid"      -> light grey background
    """
    aging_rules = [
        ("90+ Days",   RED_LIGHT,    RED_DARK),
        ("61-90 Days", AMBER_LIGHT,  AMBER_DARK),
        ("31-60 Days", YELLOW_LIGHT, YELLOW_DARK),
        ("1-30 Days",  "CCE5FF",     "004085"),
        ("Current",    GREEN_LIGHT,  GREEN_DARK),
        ("Paid",       GREY_LIGHT,   "595959"),
    ]

    def add_text_rule(ws, col_range, bucket, bg, font_color):
        rule = FormulaRule(
            formula=[f'={col_range[0]}="{bucket}"'],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(color=font_color, bold=True, name="Calibri", size=10),
        )
        ws.conditional_formatting.add(col_range, rule)

    # AR Ledger — Aging Bucket column M
    for bucket, bg, fg in aging_rules:
        add_text_rule(wb["AR Ledger"],
                      f"M{LEDGER_DATA_START}:M1000", bucket, bg, fg)

    # AP Ledger — Aging Bucket column M
    for bucket, bg, fg in aging_rules:
        add_text_rule(wb["AP Ledger"],
                      f"M{LEDGER_DATA_START}:M1000", bucket, bg, fg)

    # AR Ledger — Status column N
    status_rules = [
        ("Overdue",     RED_LIGHT,    RED_DARK),
        ("Partial",     AMBER_LIGHT,  AMBER_DARK),
        ("Outstanding", YELLOW_LIGHT, YELLOW_DARK),
        ("Paid",        GREEN_LIGHT,  GREEN_DARK),
    ]
    for status, bg, fg in status_rules:
        add_text_rule(wb["AR Ledger"],
                      f"N{LEDGER_DATA_START}:N1000", status, bg, fg)
        add_text_rule(wb["AP Ledger"],
                      f"N{LEDGER_DATA_START}:N1000", status, bg, fg)

    # Aging Report sheets — colour the bucket columns C:G
    for report_sheet in ["AR Aging Report", "AP Aging Report"]:
        ws = wb[report_sheet]
        col_rules = [
            ("G5:G20", RED_LIGHT,    RED_DARK),    # 90+ Days
            ("F5:F20", AMBER_LIGHT,  AMBER_DARK),  # 61-90
            ("E5:E20", YELLOW_LIGHT, YELLOW_DARK), # 31-60
            ("D5:D20", "CCE5FF",     "004085"),     # 1-30
            ("C5:C20", GREEN_LIGHT,  GREEN_DARK),   # Current
        ]
        for cell_range, bg, _ in col_rules:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThan", formula=["0"],
                           fill=PatternFill("solid", fgColor=bg))
            )

    print("  [ok] Conditional formatting applied")


# ─────────────────────────────────────────────────────────────────
# PHASE 5B  —  DATA VALIDATION (dropdowns)
# ─────────────────────────────────────────────────────────────────

def apply_data_validation(wb):
    """
    Adds dropdown validation to the Client Name and Vendor Name columns
    in the respective ledger sheets.
    Source lists come from the Lookup Tables sheet.

    Also adds a Status dropdown to prevent typos (though Status is formula-driven,
    the dropdown acts as a safety net for rows where users override manually).
    """
    # AR Ledger — Client Name dropdown (column D)
    ar_client_dv = DataValidation(
        type="list",
        formula1="='Lookup Tables'!$B$5:$B$9",
        allow_blank=True,
        showDropDown=False,
        error="Please select a client from the dropdown list.",
        errorTitle="Invalid Client",
        prompt="Select a client from the list",
        promptTitle="Client Name",
    )
    wb["AR Ledger"].add_data_validation(ar_client_dv)
    ar_client_dv.add(f"D{LEDGER_DATA_START}:D1000")

    # AP Ledger — Vendor Name dropdown (column D)
    ap_vendor_dv = DataValidation(
        type="list",
        formula1="='Lookup Tables'!$E$5:$E$9",
        allow_blank=True,
        showDropDown=False,
        error="Please select a vendor from the dropdown list.",
        errorTitle="Invalid Vendor",
        prompt="Select a vendor from the list",
        promptTitle="Vendor Name",
    )
    wb["AP Ledger"].add_data_validation(ap_vendor_dv)
    ap_vendor_dv.add(f"D{LEDGER_DATA_START}:D1000")

    # AR Ledger — Amount must be a positive number
    ar_amount_dv = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        error="Amount must be a positive number.",
        errorTitle="Invalid Amount",
    )
    wb["AR Ledger"].add_data_validation(ar_amount_dv)
    ar_amount_dv.add(f"F{LEDGER_DATA_START}:F1000")

    ap_amount_dv = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        error="Amount must be a positive number.",
        errorTitle="Invalid Amount",
    )
    wb["AP Ledger"].add_data_validation(ap_amount_dv)
    ap_amount_dv.add(f"F{LEDGER_DATA_START}:F1000")

    print("  [ok] Data validation applied")


# ─────────────────────────────────────────────────────────────────
# PHASE 5C  —  PRINT SETTINGS
# ─────────────────────────────────────────────────────────────────

def apply_print_settings(wb):
    """
    Sets print area, orientation, and scaling for aging report sheets
    so they print cleanly on A4 portrait.
    """
    for sheet_name in ["AR Aging Report", "AP Aging Report"]:
        ws = wb[sheet_name]

        # Print area covers titles through totals row
        ws.print_area = "A1:H12"

        ws.page_setup.orientation     = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize        = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage        = True
        ws.page_setup.fitToWidth       = 1
        ws.page_setup.fitToHeight      = 0

        # Repeat header rows 1-4 on every printed page
        ws.print_title_rows = "1:4"

        ws.oddHeader.center.text  = "&B&16AP/AR Tracker — Aging Report"
        ws.oddFooter.left.text    = "Confidential"
        ws.oddFooter.center.text  = "&P of &N"
        ws.oddFooter.right.text   = "&D"

    print("  [ok] Print settings applied")


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────

def main():
    print("\nBuilding AP/AR Tracker workbook...")
    print("=" * 44)

    print("\n[Phase 1] Workbook structure + Settings + Lookup Tables")
    wb = create_workbook()
    build_settings(wb)
    build_lookup_tables(wb)

    print("\n[Phase 2] AR Ledger + AP Ledger")
    build_ar_ledger(wb)
    build_ap_ledger(wb)

    print("\n[Phase 3] AR Aging Report + AP Aging Report")
    build_ar_aging_report(wb)
    build_ap_aging_report(wb)

    print("\n[Phase 4] Dashboard")
    build_dashboard(wb)

    print("\n[Phase 5] Conditional formatting + Data validation + Print settings")
    apply_conditional_formatting(wb)
    apply_data_validation(wb)
    apply_print_settings(wb)

    # Force Excel to recalculate all formulas when the file is opened.
    # This ensures charts and KPIs show live values on first open.
    wb.calculation.calcMode      = "auto"
    wb.calculation.fullCalcOnLoad = True

    wb.save(OUTPUT_FILE)
    print(f"\n{'=' * 44}")
    print(f"Saved -> {OUTPUT_FILE}")
    print("Open the file in Excel to review.\n")
    print("HOW TO USE:")
    print("  - Yellow cells  = input (type your data here)")
    print("  - Grey cells    = formulas (do not edit)")
    print("  - AR/AP Ledger  = add new rows below the sample data")
    print("  - Run this script again only if you want to rebuild from scratch\n")


if __name__ == "__main__":
    main()
